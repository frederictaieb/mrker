from openpyxl import Workbook
from pathlib import Path
from utils.file import delete_path
from utils.logger import get_logger
import pandas as pd

logger = get_logger(__name__)

class XlsService:
    xls_file = "data/infos/markers.xlsx"

    def __init__(self, markers=None, tracks_data=None):
        self.tracks_data = tracks_data or []
        self.markers = markers or []

    def same_len(self):
        return len(self.tracks_data) == len(self.markers)

    def min_len(self):
        return min(len(self.tracks_data), len(self.markers))
   
    def _generate_xls(self):
        wb = Workbook()
        ws = wb.active

        for i, (start, end) in enumerate(self.markers, start=1):
            ws.cell(row=i, column=1, value=start)
            ws.cell(row=i, column=2, value=end)

        for i, track_data in enumerate(self.tracks_data, start=1):
            ws.cell(row=i, column=3, value=track_data.get("id"))
            ws.cell(row=i, column=4, value=track_data.get("artist"))
            ws.cell(row=i, column=5, value=track_data.get("title"))
            ws.cell(row=i, column=6, value=track_data.get("album"))
            ws.cell(row=i, column=7, value=track_data.get("spotify_link"))
            ws.cell(row=i, column=8, value=track_data.get("duration_ms"))
            ws.cell(row=i, column=9, value=track_data.get("duration_s"))
            ws.cell(row=i, column=10, value=track_data.get("duration"))
            ws.cell(row=i, column=11, value=track_data.get("filename"))

        wb.save(self.xls_file)

    #def _generate_txt(self):
    #    with open(self.txt_file, "w", encoding="utf-8") as f:
    #        for i in range(self.min_len()):
    #            start, end = self.markers[i] 
    #            filename = self.filenames[i]
    #            f.write(f"{start}\t{end}\t{filename}\n")

    def generate(self):
        self._generate_xls()
        #if self.same_len():
        #    self._generate_txt()

    @classmethod
    def load(cls):
        if not Path(cls.xls_file).is_file():
            raise FileNotFoundError(f"Fichier introuvable : {cls.xls_file}")
        df = pd.read_excel(cls.xls_file, header=None)
        df.columns = ["start","end","id","artist","title","album","spotify_link","duration_ms","duration_s","duration","filename"]
        df = df.dropna(how="all")

        df_markers = df[["start", "end"]]
        df_markers = df_markers.dropna(subset=["start", "end"])
        markers = list(df_markers.itertuples(index=False, name=None))

        df_tracks = df[["id","artist","title","album","spotify_link","duration_ms","duration_s","duration","filename",]]
        df_tracks = df_tracks.dropna(subset=["id"])
        tracks_data = df_tracks.to_dict(orient="records")

        return tracks_data, markers



    @classmethod
    def is_generated(cls):
        return Path(cls.xls_file).is_file()

    def reset(self):
        if self.is_generated():
            delete_path(self.xls_file)
            #delete_path(self.txt_file)

            logger.info("Fichiers supprimés")
        else: 
            logger.info("Erreur")


    def __str__(self):
        return f"{self.xls_file}"

    def __repr__(self):
        return self.__str__()

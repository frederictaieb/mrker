from openpyxl import Workbook

# Créer un nouveau fichier Excel
wb = Workbook()
ws = wb.active

# Nom de la feuille
ws.title = "MaFeuille"

# Remplir les colonnes A, B, C
ws['A1'] = "Nom"
ws['B1'] = "Age"
ws['C1'] = "Ville"

ws['A2'] = "Alice"
ws['B2'] = 25
ws['C2'] = "Paris"

ws['A3'] = "Bob"
ws['B3'] = 30
ws['C3'] = "Lyon"

# Sauvegarder le fichier
wb.save("mon_fichier.xlsx")